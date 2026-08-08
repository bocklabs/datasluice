"""Streaming XLSX reader yielding Arrow ``RecordBatch`` (DATA-03, D-P4-10).

Migrated from ``datasluice.formats.xlsx``. Uses openpyxl's
``load_workbook(..., read_only=True, data_only=True)`` streaming read mode
and chunks ``iter_rows(values_only=True)`` into ``batch_size``-row
``RecordBatch`` objects via ``pa.Table.from_pylist``.

XLSX is itself a ZIP archive so openpyxl decodes it in one pass; this
reader is "streaming" in the sense that it yields batches as rows arrive
from ``iter_rows``, not that it avoids buffering the workbook. For
very wide rows the per-batch memory is not strictly bounded by byte
count (D-P4-14 acknowledges this), but open-data XLSX rows are modest
width in practice.
"""

from __future__ import annotations

from collections.abc import Iterator
from typing import Any

from datasluice.data.readers.base import BaseFormatReader
from datasluice.exceptions import FormatError


class XLSXReader(BaseFormatReader):
    """Stream an XLSX ``BinaryIO`` source into Arrow ``RecordBatch`` objects."""

    format_name = "XLSX"

    def read_batches(self, source: Any, *, batch_size: int = 65536) -> Iterator[Any]:
        """Yield ``RecordBatch`` objects by chunking openpyxl ``iter_rows``.

        Args:
            source: A binary file-like XLSX source.
            batch_size: Target rows per yielded batch.

        Raises:
            FormatError: If ``openpyxl`` / ``pyarrow`` is missing or the
                workbook is corrupt.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise FormatError("XLSX reads require 'openpyxl'. Install with: pip install datasluice[xlsx]") from exc
        try:
            import pyarrow as pa
        except ImportError as exc:
            raise FormatError(
                "Streaming reads require 'pyarrow'. Install with: pip install datasluice[streaming]"
            ) from exc

        try:
            wb = load_workbook(source, read_only=True, data_only=True)
        except Exception as exc:
            raise FormatError(f"Invalid XLSX: {exc}") from exc

        try:
            ws = wb.active
            if ws is None:
                raise FormatError("Invalid XLSX: workbook has no active worksheet")
            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                return
            raw_headers = [str(h) if h is not None else "" for h in header_row]
            # De-duplicate identical/blank header names so later cells do not
            # overwrite earlier ones inside the zipped row dicts (openpyxl can
            # emit duplicate column headers in messy real-world workbooks).
            seen: dict[str, int] = {}
            headers: list[str] = []
            for header in raw_headers:
                count = seen.get(header, 0)
                seen[header] = count + 1
                headers.append(header if count == 0 else f"{header}_{count + 1}")

            buffer: list[dict[str, Any]] = []
            for row in rows:
                buffer.append(dict(zip(headers, row, strict=False)))
                if len(buffer) >= batch_size:
                    yield _batch_from_rows(buffer, pa)
                    buffer = []
            if buffer:
                yield _batch_from_rows(buffer, pa)
        except (pa.ArrowInvalid, pa.ArrowTypeError, pa.ArrowNotImplementedError) as exc:
            raise FormatError(f"Could not coerce XLSX rows to Arrow: {exc}") from exc
        finally:
            wb.close()


def _batch_from_rows(rows: list[dict[str, Any]], pa: Any) -> Any:
    """Build a single ``RecordBatch`` from a chunk of row dicts."""
    table = pa.Table.from_pylist(rows)
    batches = table.to_batches()
    if not batches:
        return pa.RecordBatch.from_pylist(rows)
    return batches[0]
